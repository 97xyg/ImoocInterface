#coding=utf-8
import openpyxl
import sys
import os
base_path = os.getcwd()
import unittest
sys.path.append(base_path)


class HandExcel:
    def load_excel(self):
        #加载excel
        open_excel = openpyxl.load_workbook(base_path+"/Case/imooc.xlsx")
        return open_excel

    def get_sheet_data(self,index=None):
        #加载所有sheet的内容
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols):
        # 获取一个单元格的内容
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data

    def get_rows(self):
        #获取行数
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self,row):
        #
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

if __name__ == "__main__":
    handle = HandExcel()
    print(handle.get_rows_value(2))











